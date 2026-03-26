"""
Excel 讀寫處理模組
支援大夜班、小夜週、假日班的讀取和輸出
"""
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
import math
import unicodedata

from config import COLORS
from date_utils import (
    is_holiday,
    is_saturday,
    get_holiday_name,
    format_date_to_string,
    parse_date_string,
)


def get_sheet_names(filepath: str) -> List[str]:
    """取得 Excel 所有頁籤名稱"""
    wb = load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_nurses_from_excel(filepath: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """
    從 Excel 讀取護理人員資料
    
    Returns:
        [{'主值': '名字', '公休': '2/1-2/7', '大夜': '...', '小夜週': '...', '備註': '...', '_row_index': 行號}, ...]
    """
    wb = load_workbook(filepath, data_only=True)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # 找到表頭列（通常在第2或第3列）
    header_row = None
    headers = {}
    
    for row_idx in range(1, 10):  # 檢查前10列
        row_values = [cell.value for cell in ws[row_idx]]
        # 檢查是否有「主值」或「副值」欄位
        if '主值' in row_values or '副值' in row_values:
            header_row = row_idx
            for col_idx, value in enumerate(row_values, 1):
                if value:
                    headers[value] = col_idx
            break
    
    if not header_row:
        wb.close()
        return []
    
    # 讀取資料列
    nurses_data = []
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {'_row_index': row_idx}
        
        for header_name, col_idx in headers.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[header_name] = cell_value if cell_value else ''
        
        # 檢查是否有名字（主值或副值）
        name = row_data.get('主值', row_data.get('副值', ''))
        if name and str(name).strip():
            nurses_data.append(row_data)
    
    wb.close()
    return nurses_data


def find_last_assigned_from_excel(
    filepath: str,
    sheet_name: str,
    target_month: int,
    target_year: int,
) -> Optional[str]:
    """
    從 Excel 大夜欄找到上個月最後排班的人
    
    邏輯：找大夜欄中日期屬於「上個月」的最後一筆資料
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    # 計算上個月
    prev_month = target_month - 1 if target_month > 1 else 12
    prev_year = target_year if target_month > 1 else target_year - 1
    
    # 找到表頭
    header_row = None
    name_col = None
    night_col = None
    
    for row_idx in range(1, 10):
        row_values = [cell.value for cell in ws[row_idx]]
        if '主值' in row_values or '副值' in row_values:
            header_row = row_idx
            for col_idx, value in enumerate(row_values, 1):
                if value == '主值' or value == '副值':
                    name_col = col_idx
                elif value == '大夜':
                    night_col = col_idx
            break
    
    if not header_row or not name_col or not night_col:
        wb.close()
        return None

    marked_name = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        night_value = ws.cell(row=row_idx, column=night_col).value

        if not name or not night_value:
            continue

        name = str(name).strip().lstrip('*')
        value_str = str(night_value).strip()
        if '#' in value_str:
            marked_name = name

    if marked_name:
        wb.close()
        return marked_name
    
    # 找上個月最後排班的人
    last_assigned = None
    last_date = None
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        night_value = ws.cell(row=row_idx, column=night_col).value
        
        if not name or not night_value:
            continue
        
        name = str(name).strip().lstrip('*')
        night_str = str(night_value).strip()
        night_str = re.sub(r'\([^)]*\)', '', night_str).strip()
        
        # 解析大夜欄的日期（格式可能是 "2/24.2/25.2/26" 或 "2/24-2/26"）
        dates = []
        
        # 嘗試解析點分隔格式
        if '.' in night_str:
            parts = night_str.split('.')
            for part in parts:
                date = parse_date_string(part.strip(), target_year)
                if date:
                    dates.append(date)
        # 嘗試解析範圍格式
        elif '-' in night_str:
            from date_utils import parse_date_range
            dates = parse_date_range(night_str, target_year)
        else:
            date = parse_date_string(night_str, target_year)
            if date:
                dates.append(date)
        
        # 檢查是否有上個月的日期
        for date in dates:
            if date.month == prev_month:
                if last_date is None or date > last_date:
                    last_date = date
                    last_assigned = name
    
    wb.close()
    return last_assigned


def find_last_assigned_small_night_from_excel(
    filepath: str,
    sheet_name: str,
    target_month: int,
    target_year: int,
) -> Optional[str]:
    """
    從 Excel 小夜週欄找到上個月最後排班的人
    
    邏輯：找小夜週欄中日期屬於「上個月」的最後一筆資料
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    # 計算上個月
    prev_month = target_month - 1 if target_month > 1 else 12
    prev_year = target_year if target_month > 1 else target_year - 1
    
    # 找到表頭
    header_row = None
    name_col = None
    small_night_col = None
    
    for row_idx in range(1, 10):
        row_values = [cell.value for cell in ws[row_idx]]
        if '主值' in row_values or '副值' in row_values:
            header_row = row_idx
            for col_idx, value in enumerate(row_values, 1):
                if value == '主值' or value == '副值':
                    name_col = col_idx
                elif value == '小夜週':
                    small_night_col = col_idx
            break
    
    if not header_row or not name_col or not small_night_col:
        wb.close()
        return None

    marked_name = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        small_night_value = ws.cell(row=row_idx, column=small_night_col).value

        if not name or not small_night_value:
            continue

        name = str(name).strip().lstrip('*')
        value_str = str(small_night_value).strip()
        if '#' in value_str:
            marked_name = name

    if marked_name:
        wb.close()
        return marked_name
    
    # 找上個月最後排班的人
    last_assigned = None
    last_date = None
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        small_night_value = ws.cell(row=row_idx, column=small_night_col).value
        
        if not name or not small_night_value:
            continue
        
        name = str(name).strip().lstrip('*')
        small_night_str = str(small_night_value).strip()
        small_night_str = re.sub(r'\([^)]*\)', '', small_night_str).strip()
        
        # 解析小夜週欄的日期（格式通常是 "2/24-2/28"）
        dates = []
        if '-' in small_night_str:
            from date_utils import parse_date_range
            dates = parse_date_range(small_night_str, target_year)
        
        # 檢查是否有上個月的日期
        for date in dates:
            if date.month == prev_month:
                if last_date is None or date > last_date:
                    last_date = date
                    last_assigned = name
    
    wb.close()
    return last_assigned


def find_last_assigned_holiday_from_excel(
    filepath: str,
    sheet_name: str,
    target_month: int,
    target_year: int,
) -> Optional[str]:
    """
    從 Excel 假日班欄位找到上個月最後排班的人
    
    假日班欄位：現在是三個「假日」欄位
    
    邏輯：找所有假日欄位中日期屬於「上個月」的最後一筆資料
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    # 計算上個月
    prev_month = target_month - 1 if target_month > 1 else 12
    prev_year = target_year if target_month > 1 else target_year - 1
    
    # 找到表頭
    header_row = None
    name_col = None
    holiday_cols = []  # 假日欄位的索引列表
    
    for row_idx in range(1, 10):
        row_values = [cell.value for cell in ws[row_idx]]
        if '主值' in row_values or '副值' in row_values:
            header_row = row_idx
            for col_idx, value in enumerate(row_values, 1):
                if value == '主值' or value == '副值':
                    name_col = col_idx
                elif value == '假日':
                    holiday_cols.append(col_idx)
            break
    
    if not header_row or not name_col or not holiday_cols:
        wb.close()
        return None

    marked_name = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        if not name:
            continue
        name = str(name).strip().lstrip('*')
        for col_idx in holiday_cols:
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue
            cell_str = str(cell_value)
            if '#' in cell_str:
                marked_name = name

    if marked_name:
        wb.close()
        return marked_name
    
    # 找上個月最後排班的人
    last_assigned = None
    last_date = None
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        
        if not name:
            continue
        
        name = str(name).strip().lstrip('*')
        
        # 檢查所有假日欄位
        for col_idx in holiday_cols:
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            
            if not cell_value:
                continue
            
            cell_str = str(cell_value).strip()
            
            # 解析日期（格式可能是 "3/8白班" 或 "3/15白班(原3/8小夜)"）
            # 提取第一個日期
            date_match = re.search(r'(\d{1,2}/\d{1,2})', cell_str)
            if date_match:
                date = parse_date_string(date_match.group(1), target_year)
                if date and date.month == prev_month:
                    if last_date is None or date > last_date:
                        last_date = date
                        last_assigned = name
    
    wb.close()
    return last_assigned


def create_schedule_excel_multi_sheet(
    output_path: str,
    title: str,
    all_results: Dict[str, Dict],
    year: int,
    month: int,
):
    """
    建立多頁籤的排班結果 Excel
    
    all_results 格式:
    {
        '頁籤名': {
            'nurses': [NurseInfo, ...],
            'night_results': [{'nurse': NurseInfo, 'dates': [...], ...}, ...],
            'small_night_results': [...],
            'holiday_results': [...],  # 假日班結果
            'night_last_normal': '名字',
            'small_night_last_normal': '名字',
            'holiday_last_normal': '名字',
        },
        ...
    }
    """
    wb = Workbook()
    
    # 移除預設的工作表
    default_sheet = wb.active
    
    first_sheet = True
    
    for sheet_name, results in all_results.items():
        if first_sheet:
            ws = default_sheet
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        _write_schedule_sheet(
            ws, title, results, year, month, sheet_name
        )
    
    wb.save(output_path)


def _is_all_dates_other_month(text: str, year: int, month: int) -> bool:
    """
    判斷文字中的所有日期是否都屬於上個月或下個月（非當月）
    支援格式：
    - 日期範圍：4/6-4/12
    - 點分隔日期：4/2.4/3.4/4
    - 補班格式：補4/3(原4/5白班) 或 補2/27(原4/5白班)
      注意：對於補班格式，判斷的是「(原X/X班)」中的日期，而不是「補X/X」中的日期
    """
    if not text:
        return False
    
    from date_utils import parse_date_string, parse_dot_separated_dates
    
    # 計算上個月和下個月
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    # 分割成多行
    lines = text.split('\n')
    all_dates = []
    
    # 先檢查整個文字中是否有「(原X/X班)」格式
    has_original_date = False
    for line in lines:
        if not line.strip():
            continue
        original_match = re.search(r'\(原\s*(\d{1,2}/\d{1,2})', line)
        if original_match:
            has_original_date = True
            break
    
    # 如果有「(原X/X班)」格式，只提取這些日期，不提取「補X/X」中的日期
    for line in lines:
        if not line.strip():
            continue
        
        # 優先檢查補班格式中的「(原X/X班)」日期（例如：補2/27(原4/5白班) 或 補2/27\n(原4/5白班)）
        # 注意：判斷的是「(原X/X班)」中的日期，而不是「補X/X」中的日期
        original_match = re.search(r'\(原\s*(\d{1,2}/\d{1,2})', line)
        if original_match:
            date_str = original_match.group(1)
            date = parse_date_string(date_str, year)
            if date:
                all_dates.append(date)
                # 調試輸出（臨時啟用以診斷問題）
                print(f"DEBUG _is_all_dates_other_month: Found original date in line '{line}': {date_str} -> {date.month}/{date.day}, year={date.year}, next_month={next_month}, next_year={next_year}")
                continue
        
        # 如果沒有「(原X/X班)」格式，才檢查其他格式
        if not has_original_date:
            # 1. 檢查日期範圍格式（例如：4/6-4/12）
            date_range_match = re.search(r'(\d{1,2}/\d{1,2})\s*-\s*(\d{1,2}/\d{1,2})', line)
            if date_range_match:
                start_str = date_range_match.group(1)
                end_str = date_range_match.group(2)
                start_date = parse_date_string(start_str, year)
                end_date = parse_date_string(end_str, year)
                
                if start_date and end_date:
                    all_dates.extend([start_date, end_date])
                    continue
            
            # 2. 檢查點分隔日期格式（例如：4/2.4/3.4/4）
            dot_dates_match = re.search(r'(\d{1,2}/\d{1,2}(?:\.\d{1,2}/\d{1,2})+)', line)
            if dot_dates_match:
                dates_str = dot_dates_match.group(1)
                dates = parse_dot_separated_dates(dates_str, year)
                if dates:
                    all_dates.extend(dates)
                    continue
            
            # 3. 檢查「補X/X」格式（但只有在沒有「(原X/X班)」格式時才提取）
            # 注意：如果已經有「(原X/X班)」格式，就不應該提取「補X/X」中的日期
            if not has_original_date:
                makeup_match = re.search(r'補(\d{1,2}/\d{1,2})', line)
                if makeup_match:
                    date_str = makeup_match.group(1)
                    date = parse_date_string(date_str, year)
                    if date:
                        all_dates.append(date)
                        print(f"DEBUG _is_all_dates_other_month: Found makeup date in line '{line}': {date_str} -> {date.month}/{date.day}")
                        continue
    
    # 如果沒有找到任何日期，返回 False
    if not all_dates:
        print(f"DEBUG: No dates found in text: {repr(text)}")
        return False
    
    # 檢查所有日期是否都屬於下個月（非當月且非上個月）
    # 注意：只判斷下個月，上個月的補班應該用其他顏色（黑色）
    result = all(d.month == next_month and d.year == next_year for d in all_dates)
    print(f"DEBUG: dates={[(d.month, d.day) for d in all_dates]}, month={month}, next_month={next_month}, next_year={next_year}, result={result}")
    return result


def _to_inline_font(font):
    if isinstance(font, InlineFont):
        return font
    if font is None:
        return InlineFont()
    if isinstance(font, str):
        return InlineFont(rFont=font)
    color = font.color
    if isinstance(color, str):
        color = Color(rgb=color)
    return InlineFont(
        rFont=font.name,
        sz=font.size,
        b=font.bold,
        i=font.italic,
        color=color,
    )


def _build_night_rich_text(results, base_font, saturday_font, holiday_font, gray_font):
    base_font = _to_inline_font(base_font)
    saturday_font = _to_inline_font(saturday_font)
    holiday_font = _to_inline_font(holiday_font)
    gray_font = _to_inline_font(gray_font)
    parts = []
    for result_idx, result in enumerate(results):
        if result_idx > 0:
            parts.append(TextBlock(base_font, "\n"))
        dates = result.get('dates', [])
        is_next_month = result.get('is_next_month', False)
        
        # 如果是下個月的 cost，全部用灰色（包括點號）
        if is_next_month:
            for i, date in enumerate(dates):
                if i > 0:
                    parts.append(TextBlock(gray_font, "."))
                parts.append(TextBlock(gray_font, f"{date.month}/{date.day}"))
        else:
            # 當月的 cost，根據日期特性設置顏色
            for i, date in enumerate(dates):
                if i > 0:
                    parts.append(TextBlock(base_font, "."))
                if is_holiday(date):
                    font = holiday_font
                elif is_saturday(date):
                    font = saturday_font
                else:
                    font = base_font
                parts.append(TextBlock(font, f"{date.month}/{date.day}"))
        
        comp_dates = [d for d in dates if is_holiday(d) and d.weekday() < 5]
        if comp_dates:
            comp_unique = sorted({d.date(): d for d in comp_dates}.values())
            comp_text = ','.join(f"{d.month}/{d.day}" for d in comp_unique)
            comp_font = gray_font if is_next_month else base_font
            parts.append(TextBlock(comp_font, f" ({comp_text})"))

        if result.get('is_last_normal'):
            parts.append(TextBlock(base_font, "#"))
    return CellRichText(parts)


def _build_small_night_rich_text(results, base_font, holiday_font, gray_font, target_month):
    """建立小夜週的富文本格式，處理跨月灰色"""
    base_font = _to_inline_font(base_font)
    holiday_font = _to_inline_font(holiday_font)
    gray_font = _to_inline_font(gray_font)
    parts = []
    
    for result_idx, result in enumerate(results):
        if result_idx > 0:
            parts.append(TextBlock(base_font, "\n"))
        
        dates = result.get('dates', [])
        if not dates:
            continue
        
        first_date = dates[0]
        last_date = dates[-1]
        
        # 判斷整個 cost 是否為下個月
        is_next_month = result.get('is_next_month', False)
        
        if is_next_month:
            # 整個 cost 都用灰色
            parts.append(TextBlock(gray_font, f"{first_date.month}/{first_date.day}"))
            parts.append(TextBlock(gray_font, "-"))
            parts.append(TextBlock(gray_font, f"{last_date.month}/{last_date.day}"))
        else:
            # 屬於當月的 cost，逐個日期檢查是否為國定假日
            # 首日期
            if is_holiday(first_date):
                parts.append(TextBlock(holiday_font, f"{first_date.month}/{first_date.day}"))
            else:
                parts.append(TextBlock(base_font, f"{first_date.month}/{first_date.day}"))
            
            # 連接符號
            parts.append(TextBlock(base_font, "-"))
            
            # 尾日期
            if is_holiday(last_date):
                parts.append(TextBlock(holiday_font, f"{last_date.month}/{last_date.day}"))
            else:
                parts.append(TextBlock(base_font, f"{last_date.month}/{last_date.day}"))
        
        # 加上 # 標記（如果是正常輪序最後一個）
        comp_dates = [d for d in dates if is_holiday(d) and d.weekday() < 5]
        if comp_dates:
            comp_unique = sorted({d.date(): d for d in comp_dates}.values())
            comp_text = ','.join(f"{d.month}/{d.day}" for d in comp_unique)
            comp_font = gray_font if is_next_month else base_font
            parts.append(TextBlock(comp_font, f" ({comp_text})"))

        if result.get('is_last_normal'):
            parts.append(TextBlock(base_font, "#"))
    
    return CellRichText(parts)


def _get_cell_text(value) -> str:
    if isinstance(value, CellRichText):
        return ''.join(block.text for block in value)
    if value is None:
        return ''
    return str(value)


def _text_display_width(text: str) -> int:
    width = 0
    for ch in text:
        if unicodedata.east_asian_width(ch) in ('F', 'W'):
            width += 2
        else:
            width += 1
    return width


def _write_schedule_sheet(
    ws,
    title: str,
    results: Dict,
    year: int,
    month: int,
    sheet_name: str,
):
    """
    寫入單一頁籤的排班結果
    """
    from small_night_shift import format_small_night_shift_result
    
    nurses = results.get('nurses', [])
    night_results = results.get('night_results', [])
    small_night_results = results.get('small_night_results', [])
    holiday_results = results.get('holiday_results', [])
    night_identity_skipped = results.get('night_identity_skipped', set())
    small_night_identity_skipped = results.get('small_night_identity_skipped', set())
    
    # 定義欄位
    columns = [
        '主值', '公休', '備註', '大夜', '小夜週',
        '假日', '假日', '假日','假日'
    ]
    
    # 樣式定義 - 使用標楷體
    base_font_name = '標楷體'
    header_font = Font(name=base_font_name, bold=True)
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    no_wrap_center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    
    # 顏色 - 全部使用標楷體
    red_font = Font(name=base_font_name, color='FF0000')
    blue_font = Font(name=base_font_name, color='0000FF')
    green_font = Font(name=base_font_name, color='008000')
    gray_font = Font(name=base_font_name, color='A0A0A0')
    black_font = Font(name=base_font_name, color='000000')
    normal_font = Font(name=base_font_name)
    
    # 寫入標題
    ws['A1'] = title
    ws['A1'].font = Font(name=base_font_name, bold=True, size=14)
    ws['A1'].alignment = center_align
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    
    # 寫入表頭
    header_row = 2
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 建立護理人員到結果的對照表
    nurse_night_results = {}  # {name: [result, ...]}
    for result in night_results:
        name = result['nurse'].name
        if name not in nurse_night_results:
            nurse_night_results[name] = []
        nurse_night_results[name].append(result)
    
    nurse_small_night_results = {}
    for result in small_night_results:
        name = result['nurse'].name
        if name not in nurse_small_night_results:
            nurse_small_night_results[name] = []
        nurse_small_night_results[name].append(result)
    
    # 假日班結果：按輪次分配到欄位
    # 邏輯：根據 column_index 分配
    holiday_assignments = {}

    if holiday_results and nurses:
        for result in holiday_results:
            name = result.nurse.name
            new_is_comp = getattr(result, 'is_holiday_compensation', False)
            col_num = getattr(result, 'column_index', 1)
            if col_num < 1:
                col_num = 1
            if col_num > 4:
                col_num = 4
            key = (name, col_num)
            existing = holiday_assignments.get(key)
            if not existing:
                holiday_assignments[key] = [result]
            else:
                existing_is_comp = any(getattr(r, 'is_holiday_compensation', False) for r in existing)
                existing_has_normal = any(not getattr(r, 'is_holiday_compensation', False) for r in existing)
                # 同一格允許補休+正常班（擠在一起），避免重複塞相同類型
                if new_is_comp and existing_is_comp:
                    continue
                if (not new_is_comp) and existing_has_normal:
                    continue
                existing.append(result)

    # ????????
    data_row = header_row + 1
    
    for nurse in nurses:
        row_idx = data_row
        
        # 1. 主值
        cell = ws.cell(row=row_idx, column=1, value=nurse.original_name)
        cell.border = border
        cell.font = normal_font
        cell.alignment = center_align
        
        # 2. 公休
        leave_text = nurse.original_leave_text
        if leave_text:
            leave_text = re.sub(r'\s*[\u3001\uff0c,;\uff1b]\s*', '\n', leave_text)
        
        # 檢查是否包含「滿55歲跳大夜」或「55歲跳大夜」等相關文字
        has_age_55_restriction = False
        if leave_text and ('滿55歲跳大夜' in leave_text or '55歲跳大夜' in leave_text or '55歲以上跳大夜' in leave_text):
            has_age_55_restriction = True
            purple_font = Font(name='標楷體', color=COLORS['purple'])
        else:
            purple_font = None
        
        # 判斷是否所有日期都是上個月或下個月（優先級低於紫色）
        is_other_month_leave = _is_all_dates_other_month(leave_text, year, month) if leave_text else False
        
        cell = ws.cell(row=row_idx, column=2, value=leave_text)
        cell.border = border
        # 優先使用紫色（滿55歲跳大夜），其次灰色（上個月或下個月），最後預設
        if purple_font:
            cell.font = purple_font
        elif is_other_month_leave:
            cell.font = gray_font
        else:
            cell.font = normal_font
        cell.alignment = center_align
        
        # 3. 備註（全部綠色）
        # 先收集已補的待補班日期（從排班結果中）
        completed_makeup_dates = set()
        
        # 檢查大夜班補班結果
        if nurse.name in nurse_night_results:
            for result in nurse_night_results[nurse.name]:
                if result.get('is_makeup') and result.get('original_dates'):
                    # 將原始日期轉換為集合以便比較
                    original_dates_set = set(result['original_dates'])
                    completed_makeup_dates.add(('大夜', tuple(sorted(original_dates_set))))
        
        # 檢查小夜班補班結果
        if nurse.name in nurse_small_night_results:
            for result in nurse_small_night_results[nurse.name]:
                if result.get('is_makeup') and result.get('original_dates'):
                    original_dates_set = set(result['original_dates'])
                    completed_makeup_dates.add(('小夜', tuple(sorted(original_dates_set))))
        
        # 處理備註文字：移除已補的待補班資訊
        remarks_text = nurse.remarks if nurse.remarks else ''
        if remarks_text:
            # 從備註中提取所有待補班資訊
            from shift_utils import extract_pending_makeup
            pending_makeups = extract_pending_makeup(remarks_text, year)
            
            # 檢查每個待補班是否已補
            for makeup in pending_makeups:
                makeup_dates_set = set(makeup['dates'])
                makeup_key = (makeup['shift_type'], tuple(sorted(makeup_dates_set)))
                
                # 如果已補，從備註文字中移除
                if makeup_key in completed_makeup_dates:
                    from date_utils import format_dates_to_dot_string
                    first_date = min(makeup['dates'])
                    last_date = max(makeup['dates'])
                    date_range = f"{first_date.month}/{first_date.day}-{last_date.month}/{last_date.day}"
                    # 移除待補班文字（支援多種格式）
                    pattern = rf'待補班[:\s]*{re.escape(date_range)}{re.escape(makeup["shift_type"])}'
                    remarks_text = re.sub(pattern, '', remarks_text, flags=re.IGNORECASE)
                    # 移除多餘的換行和空白
                    remarks_text = re.sub(r'\n\s*\n', '\n', remarks_text).strip()
        
        remarks_parts = []
        if remarks_text.strip():
            remarks_parts.append(remarks_text.strip())
        
        # 新產生的待補班
        if nurse.new_pending_makeup:
            for makeup in nurse.new_pending_makeup:
                from date_utils import format_dates_to_dot_string
                first_date = min(makeup['dates'])
                last_date = max(makeup['dates'])
                date_range = f"{first_date.month}/{first_date.day}-{last_date.month}/{last_date.day}"
                remarks_parts.append(f"待補班:{date_range}{makeup['shift_type']}")
        
        final_remarks_text = '\n'.join(remarks_parts) if remarks_parts else ''
        green_font = Font(name='標楷體', color=COLORS['p_shift_green'])
        
        # 判斷是否所有日期都是上個月或下個月
        is_other_month_remarks = _is_all_dates_other_month(final_remarks_text, year, month) if final_remarks_text else False
        
        cell = ws.cell(row=row_idx, column=3, value=final_remarks_text)
        cell.border = border
        # 如果是上個月或下個月，使用灰色；否則使用綠色
        cell.font = gray_font if is_other_month_remarks else green_font
        cell.alignment = center_align
        
        # 取得當月身分（P1/P2/換心）
        month_key = f'{month}月'
        nurse_identity = None
        if nurse.is_transplant.get(month_key):
            nurse_identity = f'{month_key}換心'
        elif nurse.is_p1.get(month_key):
            p1_value = nurse.is_p1[month_key]
            if isinstance(p1_value, str):
                nurse_identity = f'{month_key}{p1_value}'  # 例如：3月大P1
            else:
                nurse_identity = f'{month_key}P1'
        elif nurse.is_p2.get(month_key):
            p2_value = nurse.is_p2[month_key]
            if isinstance(p2_value, str):
                nurse_identity = f'{month_key}{p2_value}'  # 例如：3月大P2
            else:
                nurse_identity = f'{month_key}P2'
        
        # 4. 大夜
        night_cell = ws.cell(row=row_idx, column=4)
        if nurse.name in nurse_night_results:
            # 有排班結果，顯示排班結果
            night_cell.value = _build_night_rich_text(
                nurse_night_results[nurse.name],
                normal_font,
                blue_font,
                red_font,
                gray_font,
            )
            night_cell.font = normal_font
        elif nurse.name in night_identity_skipped and has_age_55_restriction:
            # 有輪到但因55歲被跳過，顯示紫色標記
            night_cell.value = '滿55歲跳大夜'
            night_cell.font = purple_font
        elif nurse.name in night_identity_skipped and nurse_identity:
            # 有輪到但因身分別被跳過，顯示綠色身分
            night_cell.value = nurse_identity
            night_cell.font = green_font
        else:
            # 沒有輪到或沒有身分別，留空白
            night_cell.value = ''
            night_cell.font = normal_font
        night_cell.border = border
        night_cell.alignment = no_wrap_center_align
        
        # 5. 小夜週
        small_night_cell = ws.cell(row=row_idx, column=5)
        if nurse.name in nurse_small_night_results:
            # 有排班結果，顯示排班結果
            small_night_cell.value = _build_small_night_rich_text(
                nurse_small_night_results[nurse.name],
                normal_font,
                red_font,
                gray_font,
                month,
            )
            small_night_cell.font = normal_font
        elif nurse.name in small_night_identity_skipped and nurse_identity:
            # 有輪到但因身分別被跳過，顯示綠色身分
            small_night_cell.value = nurse_identity
            small_night_cell.font = green_font
        else:
            # 沒有輪到或沒有身分別，留空白
            small_night_cell.value = ''
            small_night_cell.font = normal_font
        small_night_cell.border = border
        small_night_cell.alignment = no_wrap_center_align
        
        # 6, 7, 8, 9. 假日班四個欄位
        for col_num in [1, 2, 3, 4]:
            col_idx = 5 + col_num  # 第 6, 7, 8, 9 欄
            holiday_cell = ws.cell(row=row_idx, column=col_idx)
            result_list = holiday_assignments.get((nurse.name, col_num))
            
            # 檢查是否有國定假日補償（優先顯示補休假資訊）
            has_holiday_compensation = False
            if result_list:
                for result in result_list:
                    if result.is_holiday_compensation:
                        has_holiday_compensation = True
                        break
            
            if has_holiday_compensation:
                # 如果有國定假日補償，顯示補休假資訊（優先於身分）
                from holiday_shift import format_holiday_shift_result
                texts = [format_holiday_shift_result(r, month) for r in result_list]
                holiday_text = '\n'.join(texts)
                holiday_cell.value = holiday_text
                
                # 判斷補班日期是否為下個月（判斷「(原X/X班)」中的日期）
                is_next_month_holiday = _is_all_dates_other_month(holiday_text, year, month) if holiday_text else False
                
                # 調試：打印判斷結果（臨時啟用以診斷問題）
                print(f"DEBUG holiday_compensation: holiday_text={repr(holiday_text)}, year={year}, month={month}, is_next_month_holiday={is_next_month_holiday}")
                
                if is_next_month_holiday:
                    # 如果原班日期是下個月，使用灰色
                    holiday_cell.font = gray_font
                else:
                    # 如果原班日期是當月，使用黑色
                    holiday_cell.font = black_font
            elif result_list:
                # 有排班結果（含跳過記錄），顯示排班結果
                from holiday_shift import format_holiday_shift_result
                texts = [format_holiday_shift_result(r, month) for r in result_list]
                holiday_text = '\n'.join(texts)
                holiday_cell.value = holiday_text

                first_result = result_list[0]
                # 判斷補班日期是否為下個月（判斷「(原X/X班)」中的日期）
                is_next_month_holiday = _is_all_dates_other_month(holiday_text, year, month) if holiday_text else False

                if first_result.slot.is_next_month or is_next_month_holiday:
                    holiday_cell.font = gray_font
                elif first_result.is_skipped:
                    # 跳過記錄：P/換心用綠色，其他用黑色
                    skip_reason = first_result.skip_reason
                    is_p_or_transplant = ('月大P1' in skip_reason or '月小P1' in skip_reason or
                                          '月大P2' in skip_reason or '月小P2' in skip_reason or
                                          '月P1' in skip_reason or '月P2' in skip_reason or
                                          '月換心' in skip_reason)
                    holiday_cell.font = green_font if is_p_or_transplant else black_font
                else:
                    holiday_cell.font = red_font
            else:
                # 沒有排班結果，留空白
                holiday_cell.value = ''
                holiday_cell.font = normal_font
            holiday_cell.border = border
            holiday_cell.alignment = center_align
        
        data_row += 1
    
    # 調整欄寬 - 確保所有欄位都能容納兩行文字
    # 主值, 公休, 備註, 大夜, 小夜週, 假日, 假日, 假日, 假日
    max_col = len(columns)
    max_row = data_row - 1
    column_widths = [0] * max_col

    # 計算每欄的最大寬度（考慮兩行文字）
    for row in range(header_row, max_row + 1):
        for col in range(1, max_col + 1):
            value = ws.cell(row=row, column=col).value
            text = _get_cell_text(value)
            lines = text.splitlines() if text else ['']
            # 計算兩行文字的最大寬度
            if len(lines) >= 2:
                # 如果有兩行或更多，取前兩行的最大寬度
                two_line_width = max(_text_display_width(line) for line in lines[:2])
                max_len = max(max(_text_display_width(line) for line in lines), two_line_width)
            else:
                # 如果只有一行，假設第二行也是同樣寬度（模擬兩行）
                single_line_width = max(_text_display_width(line) for line in lines) if lines else 0
                max_len = single_line_width * 1.2  # 稍微增加寬度以容納兩行
            if max_len > column_widths[col - 1]:
                column_widths[col - 1] = max_len

    min_widths = [8, 12, 16, 14, 12, 12, 12, 12, 12]
    max_width = None
    width_by_col = []
    for col_idx, max_len in enumerate(column_widths, 1):
        min_width = min_widths[col_idx - 1] if col_idx - 1 < len(min_widths) else 10
        # 確保寬度足夠容納兩行文字，加上一些邊距
        width = max(max_len + 2, min_width) + 5
        if max_width is not None:
            width = min(width, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        width_by_col.append(width)

    # 設定所有資料列的列高為兩行的高度（統一大小）
    two_line_height = 24 * 2  # 兩行的高度
    for row in range(header_row, max_row + 1):
        ws.row_dimensions[row].height = two_line_height

def read_cross_month_makeup_holidays_from_excel(
    filepath: str,
    sheet_name: str,
    target_month: int,
    target_year: int,
) -> Dict[str, List[Dict]]:
    """
    從 Excel 假日班欄位讀取跨月補班記錄（!開頭的）
    
    Returns:
        {
            'nurse_name': [
                {'date': datetime, 'shift_type': '白班', 'slot_type': 'sat_day', 'original_text': '!4/3白班'},
                ...
            ],
            ...
        }
    """
    from datetime import datetime
    import re
    
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    # 找到表頭
    header_row = None
    name_col = None
    holiday_cols = []  # 假日欄位的索引列表
    
    for row_idx in range(1, 10):
        row_values = [cell.value for cell in ws[row_idx]]
        if '主值' in row_values or '副值' in row_values:
            header_row = row_idx
            for col_idx, value in enumerate(row_values, 1):
                if value == '主值' or value == '副值':
                    name_col = col_idx
                elif value and '假日' in str(value):
                    holiday_cols.append(col_idx)
            break
    
    if not header_row or not name_col or not holiday_cols:
        wb.close()
        return {}
    
    # 讀取每個護理人員的跨月補班記錄
    nurse_makeups = {}
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=name_col).value
        if not name_cell:
            continue
        
        # 處理名字（移除*號）
        name = str(name_cell).lstrip('*').strip()
        if not name:
            continue
        
        # 檢查所有假日欄位
        for col_idx in holiday_cols:
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue
            
            text = str(cell_value).strip()
            
            # 檢查是否以!開頭（跨月補班標記）
            if text.startswith('!'):
                # 解析格式：!4/3白班 或 !4/3小夜
                # 移除!標記
                content = text[1:].strip()
                
                # 解析日期和班別
                # 格式：4/3白班、4/3小夜、4/3大夜
                match = re.match(r'(\d+)/(\d+)\s*(白班|小夜|大夜)', content)
                if match:
                    month = int(match.group(1))
                    day = int(match.group(2))
                    shift_type = match.group(3)
                    
                    # 只處理本月的補班記錄
                    if month == target_month:
                        try:
                            date = datetime(target_year, month, day)
                            
                            # 判斷假日班類型
                            weekday = date.weekday()
                            slot_type = None
                            if weekday == 5:  # 週六
                                if shift_type == '白班':
                                    slot_type = 'sat_day'
                                elif shift_type == '小夜':
                                    slot_type = 'sat_small_night'
                            elif weekday == 6:  # 週日
                                if shift_type == '大夜':
                                    slot_type = 'sun_night'
                                elif shift_type == '白班':
                                    slot_type = 'sun_day'
                                elif shift_type == '小夜':
                                    slot_type = 'sun_small_night'
                            else:  # 平日國定假日
                                if shift_type == '白班':
                                    slot_type = 'weekday_holiday_day'
                                elif shift_type == '小夜':
                                    slot_type = 'weekday_holiday_small_night'
                                elif shift_type == '大夜':
                                    slot_type = 'weekday_holiday_night'
                            
                            if slot_type:
                                if name not in nurse_makeups:
                                    nurse_makeups[name] = []
                                
                                nurse_makeups[name].append({
                                    'date': date,
                                    'shift_type': shift_type,
                                    'slot_type': slot_type,
                                    'original_text': text
                                })
                                
                                print(f"  讀取跨月補班: {name} {text}")
                        except ValueError:
                            # 日期無效，跳過
                            pass
    
    wb.close()
    return nurse_makeups
