"""
護理排班系統主程式
使用方式: python main.py <輸入Excel路徑>

排班順序：
1. 大夜班
2. 小夜週
3. 假日班（需要大夜和小夜的結果來檢查衝突）
"""
import sys
import os
import re
from typing import List, Dict, Tuple, Optional

import config
from date_utils import (
    roc_to_western_year,
    get_night_shift_groups_for_month,
    format_dates_to_dot_string,
)
from models import NurseInfo
from shift_utils import parse_nurse_info_from_row
from night_shift import schedule_night_shifts, format_night_shift_result
from small_night_shift import schedule_small_night_shifts, format_small_night_shift_result
from holiday_shift import (
    schedule_holiday_shifts,
    print_holiday_shift_preview,
    format_holiday_shift_result,
)
from excel_handler import (
    read_nurses_from_excel,
    get_sheet_names,
    find_last_assigned_from_excel,
    find_last_assigned_small_night_from_excel,
    find_last_assigned_holiday_from_excel,
    create_schedule_excel_multi_sheet,
)


def parse_title_year_month(filepath: str) -> Tuple[Optional[int], Optional[int]]:
    """
    從 Excel 標題解析民國年和月份
    例如: '臺北榮民總醫院護理部思源手術室115年2月值班表' -> (115, 2)
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    title = ws['A1'].value
    wb.close()
    
    if not title:
        return None, None
    
    # 匹配 XXX年X月 或 XXX年XX月
    match = re.search(r'(\d+)年(\d+)月', title)
    if match:
        roc_year = int(match.group(1))
        month = int(match.group(2))
        return roc_year, month
    
    return None, None


def find_last_assigned_index(nurses: List[NurseInfo], last_assigned_name: str) -> int:
    """找到上個月最後排班的人在列表中的索引"""
    if not last_assigned_name:
        return -1
    for idx, nurse in enumerate(nurses):
        if nurse.name == last_assigned_name:
            return idx
    return -1


def print_night_shift_preview(results: List[Dict], month: int, sheet_name: str):
    """預覽大夜班排班結果"""
    print(f"\n===== {month}月大夜班排班結果 【{sheet_name}】 =====\n")
    print(f"{'日期':<20} {'護理人員':<10} {'類型':<10} {'標記':<25}")
    print("-" * 70)

    for result in results:
        date_str = format_dates_to_dot_string(result['dates'])
        nurse_name = result['nurse'].name
        group_type = '週一～三' if result['group_type'] == 'mon-wed' else '週四～六'

        marks = []

        # 下個月標記
        if result.get('is_next_month'):
            marks.append('【下月】')

        # 補班標記
        if result.get('is_makeup'):
            marks.append('【補班】')

        if result.get('has_saturday'):
            marks.append('星期六(藍)')
        if result.get('has_holiday'):
            holiday_strs = [f"{d.month}/{d.day}" for d in result['holiday_dates']]
            marks.append(f"國定假日(紅):{','.join(holiday_strs)}")

        mark_str = ', '.join(marks) if marks else '-'

        print(f"{date_str:<20} {nurse_name:<10} {group_type:<10} {mark_str}")


def print_small_night_shift_preview(results: List[Dict], month: int, sheet_name: str):
    """預覽小夜班排班結果"""
    print(f"\n===== {month}月小夜班排班結果 【{sheet_name}】 =====\n")
    print(f"{'日期':<20} {'護理人員':<10} {'標記':<25}")
    print("-" * 55)

    for result in results:
        date_str = format_dates_to_dot_string(result['dates'])
        nurse_name = result['nurse'].name

        marks = []

        # 下個月標記
        if result.get('is_next_month'):
            marks.append('【下月】')

        # 補班標記
        if result.get('is_makeup'):
            marks.append('【補班】')

        if result.get('has_holiday'):
            holiday_strs = [f"{d.month}/{d.day}" for d in result['holiday_dates']]
            marks.append(f"國定假日(紅):{','.join(holiday_strs)}")

        mark_str = ', '.join(marks) if marks else '-'

        print(f"{date_str:<20} {nurse_name:<10} {mark_str}")


def process_sheet(
    filepath: str,
    sheet_name: str,
    year: int,
    month: int,
    last_assigned_night_name: str = None,
    last_assigned_small_night_name: str = None,
    last_assigned_holiday_name: str = None,
) -> Tuple[List[NurseInfo], List[Dict], List[Dict], List, str, str, str, set, set]:
    """
    處理單一頁籤的排班

    Returns:
        (nurses, night_shift_results, small_night_shift_results, holiday_shift_results,
         night_last_normal_name, small_night_last_normal_name, holiday_last_normal_name)
    """
    print(f"\n--- 處理頁籤: {sheet_name} ---")

    nurses_data = read_nurses_from_excel(filepath, sheet_name)

    # 解析護理人員資訊
    nurses = []
    for row_data in nurses_data:
        nurse = parse_nurse_info_from_row(row_data, row_data.get('_row_index', 0), year)
        if nurse:
            nurses.append(nurse)

    print(f"讀取到 {len(nurses)} 位護理人員")

    # 清除「預排」的國假補償記錄，保留「已確定」的記錄
    # 判斷方式：看小夜週的開始日期是否在上個月
    # - 小夜週開始日期在上個月（跨月班）→ 已確定，保留
    # - 小夜週開始日期在本月 → 預排，清除
    for nurse in nurses:
        # 從大夜欄位讀取的補償 → 全部保留（上個月的結果，已確定）
        night_comps = nurse.night_shift_holiday_compensations
        if night_comps:
            print(f"  保留 {nurse.name} 的大夜班國假補償: {[f'{d.month}/{d.day}' for d in night_comps]}")

        # 從小夜週欄位讀取的補償 → 根據小夜週開始日期判斷
        small_night_comps = nurse.small_night_shift_holiday_compensations
        if small_night_comps and nurse.small_night_shift_dates:
            # 取得小夜週的開始日期
            first_small_night_date = min(nurse.small_night_shift_dates)
            is_cross_month = first_small_night_date.month != month  # 開始日期在上個月

            if is_cross_month:
                # 跨月小夜週（如 3/30-4/3）→ 上個月排的，已確定，保留
                print(f"  保留 {nurse.name} 的跨月小夜週國假補償: {[f'{d.month}/{d.day}' for d in small_night_comps]} (小夜週開始於 {first_small_night_date.month}/{first_small_night_date.day})")
            else:
                # 本月小夜週（如 4/6-4/10）→ 預排，清除
                print(f"  清除 {nurse.name} 的小夜週預排國假補償: {[f'{d.month}/{d.day}' for d in small_night_comps]} (小夜週開始於 {first_small_night_date.month}/{first_small_night_date.day})")
                # 從 holiday_compensations_used 中移除這些記錄
                for d in small_night_comps:
                    if d in nurse.holiday_compensations_used:
                        nurse.holiday_compensations_used.remove(d)

    # 讀取跨月補班記錄（!開頭的假日班）
    print("讀取跨月補班記錄...")
    from excel_handler import read_cross_month_makeup_holidays_from_excel
    cross_month_makeups = read_cross_month_makeup_holidays_from_excel(filepath, sheet_name, month, year)
    
    # 填充到每個護理人員
    for nurse in nurses:
        if nurse.name in cross_month_makeups:
            nurse.previous_month_cross_month_makeup_holidays = cross_month_makeups[nurse.name]

    # 如果沒有護理人員，直接返回空結果
    if not nurses:
        print(f"  頁籤 {sheet_name} 沒有護理人員資料，跳過")
        return [], [], [], [], '', '', '', set(), set()

    # ===========================
    # 1. 大夜班排班（第一步）
    # ===========================
    # 如果沒有指定上月最後排班者，從 Excel 大夜欄自動讀取
    if not last_assigned_night_name:
        last_assigned_night_name = find_last_assigned_from_excel(filepath, sheet_name, month, year)
        if last_assigned_night_name:
            print(f"從 Excel 讀取到大夜班上個月最後排班: {last_assigned_night_name}")

    # 找到上個月最後排班的人的索引
    last_night_index = find_last_assigned_index(nurses, last_assigned_night_name)
    if last_assigned_night_name and last_night_index >= 0:
        print(f"大夜班上個月最後排班: {last_assigned_night_name} (索引: {last_night_index})")
    elif last_assigned_night_name:
        print(f"警告: 找不到 {last_assigned_night_name}，從第一個人開始排")

    # 執行大夜班排班
    print("開始大夜班排班...")
    night_results, night_last_normal_name, night_identity_skipped = schedule_night_shifts(
        nurses, year, month, last_night_index
    )

    # 預覽大夜班結果
    print_night_shift_preview(night_results, month, sheet_name)

    # ===========================
    # 2. 小夜班排班（第二步）
    # ===========================
    # 如果沒有指定上月最後排班者，從 Excel 小夜週欄自動讀取
    if not last_assigned_small_night_name:
        last_assigned_small_night_name = find_last_assigned_small_night_from_excel(filepath, sheet_name, month, year)
        if last_assigned_small_night_name:
            print(f"從 Excel 讀取到小夜班上個月最後排班: {last_assigned_small_night_name}")

    # 找到上個月最後排班的人的索引
    last_small_night_index = find_last_assigned_index(nurses, last_assigned_small_night_name)
    if last_assigned_small_night_name and last_small_night_index >= 0:
        print(f"小夜班上個月最後排班: {last_assigned_small_night_name} (索引: {last_small_night_index})")
    elif last_assigned_small_night_name:
        print(f"警告: 找不到 {last_assigned_small_night_name}，從第一個人開始排")

    # 執行小夜班排班
    print("開始小夜班排班...")
    small_night_results, small_night_last_normal_name, small_night_identity_skipped = schedule_small_night_shifts(
        nurses, year, month, last_small_night_index
    )

    # 預覽小夜班結果
    print_small_night_shift_preview(small_night_results, month, sheet_name)

    # ===========================
    # 3. 假日班排班（第三步，需要大夜和小夜的結果）
    # ===========================
    # 如果沒有指定上月最後排班者，從 Excel 假日班欄位自動讀取
    if not last_assigned_holiday_name:
        last_assigned_holiday_name = find_last_assigned_holiday_from_excel(filepath, sheet_name, month, year)
        if last_assigned_holiday_name:
            print(f"從 Excel 讀取到假日班上個月最後排班: {last_assigned_holiday_name}")

    # 找到上個月最後排班的人的索引
    last_holiday_index = find_last_assigned_index(nurses, last_assigned_holiday_name)
    if last_assigned_holiday_name and last_holiday_index >= 0:
        print(f"假日班上個月最後排班: {last_assigned_holiday_name} (索引: {last_holiday_index})")
    elif last_assigned_holiday_name:
        print(f"警告: 找不到 {last_assigned_holiday_name}，從第一個人開始排")

    # 執行假日班排班
    print("開始假日班排班...")
    holiday_results, holiday_last_normal_name = schedule_holiday_shifts(
        nurses, year, month, last_holiday_index
    )

    # 預覽假日班結果
    print_holiday_shift_preview(holiday_results, month, sheet_name)

    return (
        nurses,
        night_results,
        small_night_results,
        holiday_results,
        night_last_normal_name,
        small_night_last_normal_name,
        holiday_last_normal_name,
        night_identity_skipped,
        small_night_identity_skipped,
    )


def main():
    """主程式"""
    input_file = None

    # 解析命令列參數
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]

    if not input_file or not os.path.exists(input_file):
        print("使用方式: python main.py <輸入Excel路徑>")
        print("範例: python main.py 115年3月值班表.xlsx")
        return

    # 從 Excel 標題自動讀取年份和月份
    roc_year, month = parse_title_year_month(input_file)
    if roc_year is None or month is None:
        print("錯誤: 無法從 Excel 標題解析年份和月份")
        print("請確認標題格式為: XXX年X月值班表")
        return

    # 轉換年份
    year = roc_to_western_year(roc_year)

    print(f"護理排班系統")
    print(f"目標月份: 民國 {roc_year} 年 {month} 月 (西元 {year} 年)")
    print(f"國定假日數量: {len(config.HOLIDAYS_2026)} 天")
    print(f"\n讀取檔案: {input_file}")

    # 取得所有頁籤
    sheet_names = get_sheet_names(input_file)
    print(f"找到 {len(sheet_names)} 個頁籤: {sheet_names}")

    # 處理每個頁籤
    all_results = {}

    for sheet_name in sheet_names:
        # process_sheet 會自動從 Excel 讀取上月最後排班者
        (
            nurses,
            night_results,
            small_night_results,
            holiday_results,
            night_last_normal,
            small_night_last_normal,
            holiday_last_normal,
            night_identity_skipped,
            small_night_identity_skipped,
        ) = process_sheet(
            input_file,
            sheet_name,
            year,
            month,
        )

        all_results[sheet_name] = {
            'nurses': nurses,
            'night_results': night_results,
            'small_night_results': small_night_results,
            'holiday_results': holiday_results,
            'night_last_normal': night_last_normal,
            'small_night_last_normal': small_night_last_normal,
            'holiday_last_normal': holiday_last_normal,
            'night_identity_skipped': night_identity_skipped,
            'small_night_identity_skipped': small_night_identity_skipped,
        }

    # 輸出到新的 Excel（多頁籤）
    output_file = f"排班結果_{roc_year}年{month}月.xlsx"
    title = f"臺北榮民總醫院護理部思源手術室{roc_year}年{month}月值班表"

    create_schedule_excel_multi_sheet(output_file, title, all_results, year, month)
    print(f"\n排班結果已輸出到: {output_file}")

    # 輸出摘要
    print("\n" + "=" * 70)
    print("排班摘要")
    print("=" * 70)
    
    for sheet_name, results in all_results.items():
        print(f"\n【{sheet_name}】")
        print(f"  護理人員數: {len(results['nurses'])}")
        print(f"  大夜班次數: {len(results['night_results'])}")
        print(f"  小夜週次數: {len(results['small_night_results'])}")
        print(f"  假日班次數: {len(results['holiday_results'])}")
        print(f"  定位點:")
        print(f"    大夜: {results['night_last_normal']}")
        print(f"    小夜週: {results['small_night_last_normal']}")
        print(f"    假日班: {results['holiday_last_normal']}")


if __name__ == '__main__':
    main()